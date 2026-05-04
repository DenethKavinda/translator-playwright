from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(__file__).resolve().parent / \
    "Assignment 1 - Test cases.xlsx"
SHEET_NAME = " Test cases"
TEMPLATE_ROW = 2


CASES = [
    {"input": "oya ai da mehema katha karanne?",
        "expected": "ඔයා ඇයි ද මෙහෙම කතා කරන්නේ?"},
    {"input": "ehema thiyenne naththam kiyanna.",
        "expected": "එහෙම තියෙන්නේ නැත්තම් කියන්න."},
    {"input": "eka email eken wath send karanna.",
        "expected": "එක email එකෙන්වත් send කරන්න."},
    {"input": "photo eka crop karala archive eke daanna.",
        "expected": "photo එක crop කරලා archive එකේ දාන්න."},
    {"input": "suba udasak, mage sahodarayaa!",
        "expected": "සුබ උදෑසනක්, මගේ සහෝදරයා!"},
    {"input": "suba dawasak wewa yaluwo!", "expected": "සුබ දවසක් වේවා යාළුවෝ!"},
    {"input": "tika paara innako, mama balanna.",
        "expected": "ටිකක් පාර ඉන්නකෝ, මම බලන්නම්."},
    {"input": "karunakarala queue eke idala mata help ekak denna.",
        "expected": "කරුණාකරලා queue එකේ ඉඳලා මට help එකක් දෙන්න."},
    {"input": "hari, mama tikak witarai karanne.",
        "expected": "හරි, මම ටිකක් විතරයි කරන්නේ."},
    {"input": "ane epa, mata dena eka yanna.",
        "expected": "අනේ එපා, මට දෙන එක යන්න."},
    {"input": "hari hari enna, enna.", "expected": "හරි හරි එන්න, එන්න."},
    {"input": "podi podi details tika update karanna.",
        "expected": "පොඩි පොඩි details ටික update කරන්න."},
    {"input": "mata answer eka denna... quick?",
        "expected": "මට answer එක දෙන්න... quick?"},
    {"input": "oya yannada?! epa!", "expected": "ඔයා යනවද?! එපා!"},
    {"input": "api ape rataata adarei.", "expected": "අපි අපේ රටට ආදරෙයි."},
    {"input": "man oyata passe katha krnam.",
        "expected": "මං ඔයාට පස්සේ කතා කරන්නම්."},
    {"input": "mama ada lunch ekak ganna office ekata yanawa.",
        "expected": "මම අද lunch එකක් ගන්න office එකට යනවා."},
    {"input": "eyage report eka edit karanna oni excel walin.",
        "expected": "එයාගේ report එක edit කරන්න ඕනි Excel වලින්."},
    {"input": "api den all hands meeting eken passe coffee gannam.",
        "expected": "අපි දැන් all hands meeting එකෙන් පස්සේ coffee ගන්නම්."},
    {"input": "eka hari, but we can fix it later.",
        "expected": "එක හරි, but we can fix it later."},
    {"input": "wifi speed eka tikak slow wage.",
        "expected": "WiFi speed එක ටිකක් slow වගේ."},
    {"input": "backup file eka cloud eken restore karanna.",
        "expected": "backup file එක cloud එකෙන් restore කරන්න."},
    {"input": "mama Instagram story ekak damma.",
        "expected": "මම Instagram story එකක් දැම්මා."},
    {"input": "team chat eka Discord walin karamu.",
        "expected": "team chat එක Discord වලින් කරමු."},
    {"input": "manager kiyala SOP eka ASAP nathuwa ewanna epa.",
        "expected": "manager කියලා SOP එක ASAP නැතුව එවන්න එපා."},
    {"input": "mama ETA eka dannam kiwwa.", "expected": "මම ETA එක දන්නම් කිව්වා."},
    {"input": "exam ekata prep tikak thiyenawa.",
        "expected": "exam එකට prep ටිකක් තියෙනවා."},
    {"input": "lab eke demo eka awith thama set wenne.",
        "expected": "lab එකේ demo එක ඇවිත් තාම set වෙන්නේ."},
    {"input": "api rajagiriya side eken turn wenamu.",
        "expected": "අපි Rajagiriya side එකෙන් turn වෙමු."},
    {"input": "evening eke Galle road side ekata yana eka lesi.",
        "expected": "evening එකේ Galle road side එකට යන එක ලේසි."},
    {"input": "Nimal awith lesson eka start kala.",
        "expected": "Nimal ඇවිත් lesson එක start කලා."},
    {"input": "Tharushi kiwwa eya hariyata nodanne kiyala.",
        "expected": "Tharushi කිව්වා එයා හරියට නොදන්නේ කියලා."},
    {"input": "2nd round eke 3 girls and 4 boys hitiya.",
        "expected": "2nd round එකේ 3 girls and 4 boys හිටියා."},
    {"input": "mama 10k wage steps dawasata wadi kala.",
        "expected": "මම 10k වගේ steps දවසට වැඩි කලා."},
    {"input": "mata EUR 1200 ekak transfer karanna ona.",
        "expected": "මට EUR 1200 එකක් transfer කරන්න ඕන."},
    {"input": "land eka Rs. 3,250,000 ta gaththa.",
        "expected": "land එක Rs. 3,250,000 ට ගත්තා."},
    {"input": "train eka 6:15am pamanak delay una.",
        "expected": "train එක 6:15am පමණක් delay උනා."},
    {"input": "meeting eka 21.45 ta start karanawa.",
        "expected": "meeting එක 21.45 ට start කරනවා."},
    {"input": "2026-12-24ta event eka danna.",
        "expected": "2026-12-24ට event එක දාන්න."},
    {"input": "mama next Monday call karannam.",
        "expected": "මම next Monday call කරන්නම්."},
    {"input": "mama 5kg rice gaththa.", "expected": "මම 5kg rice ගත්තා."},
    {"input": "eya 12km wage duwanawa.", "expected": "එයා 12km වගේ දුවනවා."},
    {"input": "mokada bro, scene eka flex ne.",
        "expected": "මොකද bro, scene එක flex නේ."},
    {"input": "uba hitapu widihata nam ekama vibe.",
        "expected": "උඹ හිතපු විදිහට නම් එකම vibe."},
    {
        "input": "meka balanna: https://example.lk/search?q=1, screenshot eka ganin, result eka copy karala report eke daanna, email anna@sample.net walin reply karanna, team group eke pin karanna, manager ta update karanna, follow up call eka arrange karanna, and final version eka lecturer ta send karanna before Friday.",
        "expected": "මේක බලන්න: https://example.lk/search?q=1, screenshot එක ගන්න, result එක copy කරලා report එකේ දාන්න, email anna@sample.net වලින් reply කරන්න, team group එකේ pin කරන්න, manager ට update කරන්න, follow up call එක arrange කරන්න, and final version එක lecturer ට send කරන්න before Friday.",
    },
    {"input": "email eka thiyanawa anna@sample.net, check karanna.",
        "expected": "email එක තියෙනවා anna@sample.net, check කරන්න."},
    {
        "input": "mata me denuma hari giya 😅, but still maru scene ekak thama thiyenawa, because client ekata thawa tikak explain karanna one, screenshots tika send karanna, and then feedback eka ganna one before you go home.",
        "expected": "මට මේ දැනුම හරි ගියා 😅, but still මරු scene එකක් තාම තියෙනවා, because client එකට තව ටිකක් explain කරන්න ඕන, screenshots ටික send කරන්න, and then feedback එක ගන්න ඕන before you go home.",
    },
    {
        "input": "oya karapu wada eka supiri 😎, eath office ekata gihin report eka denna, slide deck eka update karanna, tomorrow morning 8:30am call ekata ready wenna, otherwise boss eka apahu ahai.",
        "expected": "ඔයා කරපු වැඩ එක සුපිරි 😎, ඒත් office එකට ගිහින් report එක දෙන්න, slide deck එක update කරන්න, tomorrow morning 8:30am call එකට ready වෙන්න, otherwise boss එක ආපහු අහයි.",
    },
    {
        "input": "api hitanne thawa tikak wait karanna one, because QA team eka thawa bug ekak verify karanawa, after that final build eka share karala, git link eka and documentation eka samaga email karanna, nathnam marking eken issue ekak enna puluwan.",
        "expected": "අපි හිතන්නේ තව ටිකක් wait කරන්න ඕන, because QA team එක තව bug එකක් verify කරනවා, after that final build එක share කරලා, git link එක and documentation එක සමග email කරන්න, නැත්නම් marking එකෙන් issue එකක් එන්න පුළුවන්.",
    },
    {
        "input": "uba denna one summary eka issara, then I can paste it into the workbook, send it to the lecturer, and keep a copy in the folder, because if the file changes after this we need a clear version for comparison and review.",
        "expected": "උඹ දෙන්න ඕන summary එක ඉස්සර, then I can paste it into the workbook, send it to the lecturer, and keep a copy in the folder, because if the file changes after this we need a clear version for comparison and review.",
    },
]


def infer_length_type(text: str) -> str:
    length = len(text)
    if length <= 30:
        return "S"
    if length <= 299:
        return "M"
    return "L"


def copy_template_row(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def seed_workbook(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active

    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    required_rows = 1 + len(CASES)
    if ws.max_row > required_rows:
        ws.delete_rows(required_rows + 1, ws.max_row - required_rows)

    for index, case in enumerate(CASES, start=2):
        copy_template_row(ws, TEMPLATE_ROW, index, 6)
        ws.cell(index, 1).value = f"Neg_{index - 1:04d}"
        ws.cell(index, 2).value = infer_length_type(case["input"])
        ws.cell(index, 3).value = case["input"]
        ws.cell(index, 4).value = case["expected"]
        ws.cell(index, 5).value = None
        ws.cell(index, 6).value = None

    wb.save(workbook_path)
    print(f"Seeded {len(CASES)} test cases into {workbook_path}")


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")
    seed_workbook(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
