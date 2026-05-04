import re
from pathlib import Path
import openpyxl

WB = Path("Assignment 1 - Test cases.xlsx")
SHEET = " Test cases"

url_re = re.compile(r"https?://|www\.|\w+\.[a-z]{2,}")
email_re = re.compile(r"\S+@\S+\.[A-Za-z]{2,}")
time_re = re.compile(r"\b\d{1,2}:\d{2}\b")
date_re = re.compile(r"\b\d{4}-\d{2}-\d{2}\b")
currency_re = re.compile(r"\b(Rs\.?|EUR|\$)\b", re.IGNORECASE)
number_re = re.compile(r"\d+")
emoji_chars = set(["😅", "😎"])  # known emojis in this dataset


def normalize(s):
    return "" if s is None else str(s).strip()


def classify_input(text):
    t = normalize(text)
    types = []
    evidence = []
    if not t:
        return "", ""
    if url_re.search(t):
        types.append("URL")
        evidence.append("Contains URL/website")
    if email_re.search(t):
        types.append("Email")
        evidence.append("Contains email address")
    if any(e in t for e in emoji_chars):
        types.append("Emoji")
        found = ",".join([e for e in emoji_chars if e in t])
        evidence.append(f"Contains emoji: {found}")
    if currency_re.search(t):
        types.append("Currency")
        evidence.append("Currency abbreviation/format present")
    if time_re.search(t) or date_re.search(t):
        types.append("Time/Date")
        evidence.append("Contains time/date pattern")
    if number_re.search(t) and not currency_re.search(t):
        types.append("Number")
        evidence.append("Numeric digits present")
    # Mixed English/technical words
    english_tokens = ["email", "instagram", "discord", "wifi", "backup",
                      "SOP", "ETA", "call", "report", "slide", "Git", "github", "link"]
    if any(tok.lower() in t.lower() for tok in english_tokens):
        types.append("Mixed English/Loanwords")
        evidence.append("Contains English terms or technical loanwords")
    # punctuation/commands/list
    if any(p in t for p in [":", ",", "...", "-", "/", ";"]):
        types.append("Punctuation/Multiple commands")
        evidence.append(
            "Punctuation or multiple command-like segments present")
    # length based
    if len(t) > 120:
        types.append("Long/Complex")
        evidence.append("Long or multi-action instruction")

    # remove duplicates preserving order
    seen = set()
    types_clean = []
    for x in types:
        if x not in seen:
            seen.add(x)
            types_clean.append(x)
    return ", ".join(types_clean), "; ".join(evidence)


if __name__ == '__main__':
    wb = openpyxl.load_workbook(WB)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active

    # find header row heuristically
    def _normalize_header(v):
        if v is None:
            return ""
        return re.sub(r"[^a-z0-9]+", "", str(v).strip().lower())

    header_row = 1
    max_scan = min(30, ws.max_row or 30)
    for r in range(1, max_scan + 1):
        vals = [c.value for c in ws[r]]
        norms = {_normalize_header(v) for v in vals if isinstance(v, str)}
        if "input" in norms and ("expectedoutput" in norms or "expected" in norms):
            header_row = r
            break

    # build header values
    header_values = [ws.cell(row=header_row, column=c).value for c in range(
        1, (ws.max_column or 1) + 1)]

    def find_col_index(candidates):
        for i, v in enumerate(header_values, start=1):
            if v is None:
                continue
            n = _normalize_header(v)
            for c in candidates:
                if _normalize_header(c) == n:
                    return i
        # fallback: match by substring
        for i, v in enumerate(header_values, start=1):
            if v is None:
                continue
            for c in candidates:
                if _normalize_header(c) in _normalize_header(v) or _normalize_header(v) in _normalize_header(c):
                    return i
        return None

    input_candidates = ["Singlish", "Input", "Singlish Input",
                        "Test Input", "Source", "Sentence", "Text"]
    status_candidates = ["Status", "Result", "Pass/Fail", "Pass Fail"]

    input_idx = find_col_index(input_candidates)
    status_idx = find_col_index(status_candidates)
    if not input_idx:
        raise SystemExit("Could not resolve input column")
    if not status_idx:
        # if status not found, append at end
        status_idx = ws.max_column

    # read inputs before modifying sheet
    data_rows = list(range(header_row + 1, (ws.max_row or 0) + 1))
    inputs = [normalize(ws.cell(r, input_idx).value) for r in data_rows]

    # insert two columns immediately after status_idx
    insert_at = status_idx + 1
    ws.insert_cols(insert_at, amount=2)

    col_a = insert_at
    col_b = insert_at + 1
    ws.cell(header_row, col_a).value = "Singlish input types covered"
    ws.cell(header_row, col_b).value = "Evidence or rationale for the input type covered"

    # populate rows
    for idx, r in enumerate(data_rows):
        row = r
        text = inputs[idx]
        types, evidence = classify_input(text)
        ws.cell(row, col_a).value = types
        ws.cell(row, col_b).value = evidence

    wb.save(WB)
    print(
        f"Inserted columns at {col_a},{col_b} and populated {len(data_rows)} rows")
