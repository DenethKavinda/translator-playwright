import re
from pathlib import Path
import openpyxl

WB = Path("Assignment 1 - Test cases.xlsx")
SHEET = " Test cases"

if __name__ == '__main__':
    wb = openpyxl.load_workbook(WB)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active

    # find header row heuristically
    def _normalize_header(v):
        if v is None:
            return ""
        return re.sub(r"[^a-z0-9]+", "", str(v).strip().lower())

    header_row = 1
    max_scan = min(30, ws.max_row or 30)
    for r in range(1, max_scan + 1):
        vals = [c.value for c in ws[r]]
        norms = {_normalize_header(v) for v in vals if isinstance(v, str)}
        if "input" in norms and ("expectedoutput" in norms or "expected" in norms):
            header_row = r
            break

    # build header values
    header_values = [ws.cell(row=header_row, column=c).value for c in range(
        1, (ws.max_column or 1) + 1)]

    # find and delete the columns
    columns_to_delete = []
    for i, v in enumerate(header_values, start=1):
        if v is None:
            continue
        n = _normalize_header(v)
        if n in ["singlishinputtypescovered", "evidenceorrationalefortheinputtypecovered"]:
            columns_to_delete.append(i)

    # delete columns in reverse order to maintain correct indices
    for col_idx in sorted(columns_to_delete, reverse=True):
        ws.delete_cols(col_idx, amount=1)
        print(f"Deleted column {col_idx}: {header_values[col_idx - 1]}")

    wb.save(WB)
    print(f"Removed {len(columns_to_delete)} columns from the workbook")
