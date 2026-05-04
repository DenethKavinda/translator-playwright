import openpyxl
wb = openpyxl.load_workbook('Assignment 1 - Test cases.xlsx')
ws = wb[' Test cases']
header = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
print('HEADER:', header)
for r in range(2, 7):
    print(r, ws.cell(r, 7).value, ' | ', ws.cell(r, 8).value)
